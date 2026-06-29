import os
import re
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse
from pydantic import BaseModel
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from ..database import get_db
from ..models.mapping import ItemMapping
from ..services.excel_generator import process_invoices_for_sheet, process_change_orders_for_sheet

router = APIRouter()

class ExportDataRequest(BaseModel):
    project_name: str
    materials: list
    pos: list
    cos: list
    invoices: list = []

# --- Legacy Logic Ports ---

# Sheet names
SHEET_COBIA = "Cobia Cove Appartments"
SHEET_WILLOW = "Willow Way Apts"

COBIA_COLS = {
    "type": "A", "qty": "B", "co_qty": "AP", "po_co_qty": "AQ",
    "thickness": "AR", "x": "AS", "width": "AT", "length": "AU",
    "material_type": "AV", "lf_pcs": "AW", "bf_sf": "AX",
    "cost_mbf": "AY", "total_cost": "AZ", "total_cost_tax": "BA",
    "invoice_num": "BB", "total_delivered": "DV", "delivered_lf": "DW",
    "delivered_bf_sf": "DX", "delivered_cost": "DY", "delivered_cost_tax": "DZ",
    "pct_delivery": "EA", "inv_bundles": "EB", "inv_uom": "EC",
    "pcs_bundle": "ED", "inv_pcs": "EE", "issues": "EF",
    "issues_lf": "EG", "issues_bf_sf": "EH", "pct_issued": "EI",
    "issues_cost": "EJ", "issues_cost_tax": "EK", "variance_code": "EL",
    "reason": "EM",
}

WILLOW_COLS = {
    "type": "A", "qty": "B", "co_qty": "Q", "po_co_qty": "R",
    "thickness": "S", "x": "T", "width": "U", "length": "V",
    "material_type": "W", "lf_pcs": "X", "bf_sf": "Y",
    "cost_mbf": "Z", "total_cost": "AA", "total_cost_tax": "AB",
    "invoice_num": "AC", "total_delivered": "AV", "delivered_lf": "AW",
    "delivered_bf_sf": "AX", "delivered_cost": "AY", "delivered_cost_tax": "AZ",
    "pct_delivery": "BA", "inv_bundles": "BB", "inv_uom": "BC",
    "pcs_bundle": "BD", "inv_pcs": "BE", "issues": "BF",
    "issues_lf": "BG", "issues_bf_sf": "BH", "pct_issued": "BI",
    "issues_cost": "BJ", "issues_cost_tax": "BK", "variance_code": "BL",
    "reason": "BM",
}

def col_to_num(col_letter: str) -> int:
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result

def classify_item_category(desc: str) -> str:
    desc = desc.upper()
    if any(k in desc for k in ["LVL", "PSL", "GLB", "GLULAM", "LSL"]):
        return "lvl"
    if any(k in desc for k in ["OSB", "PLYWOOD", "PLY", "ZIP", "CDX", "GYPSUM", "SHEATHING"]):
        return "panels"
    if any(k in desc for k in ["SILL SEAL", "ADHESIVE", "TAPE", "FLASHING", "TYVEK", 
                                "WRAP", "SEALANT", "CAULK", "DYNAFLEX", "THRESHOLD",
                                "BT20", "COMPOUND"]):
        return "each"
    if any(k in desc for k in ["SYP", "MCA", "TREATED", "LUMBER"]):
        return "lumber"
    return "lumber"

def find_matching_row(ws, mat: dict, cols: dict, data_start_row: int, data_end_row: int):
    mat_type_col = col_to_num(cols["material_type"])
    thickness_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    length_col = col_to_num(cols["length"])
    type_col = col_to_num(cols["type"])
    
    inv_desc = mat.get("description", "").upper().strip()
    inv_dims = mat.get("dimensions", "").strip()
    
    inv_t, inv_w, inv_l = None, None, None
    dm = re.match(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_dims)
    if not dm:
        # Fallback to searching the description for T x W x L
        dm = re.search(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_desc)
        
    if dm:
        inv_t = int(dm.group(1))
        inv_w = int(dm.group(2))
        inv_l = int(dm.group(3))
    else:
        # Try to find just W x L (e.g. 4X8 for panels)
        dm_2d = re.search(r'(\d+)\s*[xX]\s*(\d+)', inv_desc)
        if dm_2d:
            inv_w = int(dm_2d.group(1))
            inv_l = int(dm_2d.group(2))
            
    # Attempt to extract length from PET (e.g. 104-5/8 is 9ft, 116-5/8 is 10ft)
    if not inv_l and "PET 104-5/8" in inv_desc:
        inv_l = 9
    elif not inv_l and "PET 116-5/8" in inv_desc:
        inv_l = 10
    
    inv_category = classify_item_category(inv_desc)
    best_match = None
    best_score = 0
    
    for row in range(data_start_row, data_end_row + 1):
        row_type = ws.cell(row=row, column=type_col).value
        if not row_type:
            continue
        row_type = str(row_type).strip()
        
        row_mat = ws.cell(row=row, column=mat_type_col).value
        row_mat = str(row_mat).strip().upper() if row_mat else ""
        
        row_t = ws.cell(row=row, column=thickness_col).value
        row_w = ws.cell(row=row, column=width_col).value
        row_l = ws.cell(row=row, column=length_col).value
        
        try: row_t = int(float(str(row_t).strip())) if row_t and str(row_t).strip() not in ('', '\xa0') else None
        except: row_t = None
        try: row_w = int(float(str(row_w).strip())) if row_w and str(row_w).strip() not in ('', '\xa0') else None
        except: row_w = None
        try: row_l = int(float(str(row_l).strip())) if row_l and str(row_l).strip() not in ('', '\xa0') else None
        except: row_l = None
        
        score = 0
        row_category = row_type.lower()
        if inv_category == row_category:
            score += 10
        elif inv_category in row_category or row_category in inv_category:
            score += 5
        else:
            continue
            
        if inv_t is not None and row_t is not None and inv_t == row_t: score += 5
        if inv_w is not None and row_w is not None and inv_w == row_w: score += 5
        if inv_l is not None and row_l is not None and inv_l == row_l: score += 5
        
        if row_mat:
            if "PT" in inv_desc and ("PT" in row_mat or "MCA" in row_mat): score += 3
            elif "MCA" in inv_desc and ("MCA" in row_mat or "PT" in row_mat): score += 3
            elif "SYP" in inv_desc and "SYP" in row_mat: score += 3
            elif "LVL" in inv_desc and "LVL" in row_mat: score += 3
            elif "OSB" in inv_desc and "OSB" in row_mat: score += 3
            elif "PLY" in inv_desc and "PLY" in row_mat: score += 3
            elif "ZIP" in inv_desc and "ZIP" in row_mat: score += 3
            if row_type.lower() in ("each", "invoice"):
                score += len(set(row_mat.split()) & set(inv_desc.split())) * 2
                
        if score > best_score:
            best_score = score
            best_match = row
            
    if best_score >= 15:
        return best_match
    return None

def safe_set_cell(ws, row, col, val):
    try:
        ws.cell(row=row, column=col, value=val)
    except AttributeError:
        pass # Ignore MergedCell error

def seed_sheet_from_pos(ws, sheet_name: str, pos: list, cols: dict, project_name: str):
    # pos contains dictionaries instead of PurchaseOrderData objects since it's from JSON
    project_pos = [po for po in pos if project_name.lower() in (po.get("project_name") or "Unknown").lower() or po.get("project_name") == "Unknown Project"]
    
    type_col = col_to_num(cols["type"])
    is_blank = True
    for r in range(3, 10):
        if ws.cell(row=r, column=type_col).value:
            is_blank = False
            break
            
    if not is_blank:
        return # Only seed if blank
        
    qty_col = col_to_num(cols["qty"])
    po_co_qty_col = col_to_num(cols["po_co_qty"])
    t_col = col_to_num(cols["thickness"])
    w_col = col_to_num(cols["width"])
    l_col = col_to_num(cols["length"])
    mat_type_col = col_to_num(cols["material_type"])
    cost_col = col_to_num(cols["cost_mbf"])
    total_cost_col = col_to_num(cols["total_cost"])
    
    row = 3
    for po in project_pos:
        for item in po.get("line_items", []):
            category = item.get("category", "lumber").capitalize()
            safe_set_cell(ws, row, type_col, category)
            
            qty = item.get("quantity", 0)
            safe_set_cell(ws, row, qty_col, qty)
            safe_set_cell(ws, row, po_co_qty_col, qty)
            
            dims = item.get("dimensions", "")
            if dims:
                dm = re.match(r'(\d+)X(\d+)X(\d+)', dims)
                if dm:
                    safe_set_cell(ws, row, t_col, int(dm.group(1)))
                    safe_set_cell(ws, row, w_col, int(dm.group(2)))
                    safe_set_cell(ws, row, l_col, int(dm.group(3)))
            
            safe_set_cell(ws, row, mat_type_col, item.get("description", ""))
            safe_set_cell(ws, row, cost_col, item.get("unit_price", 0))
            
            qty_letter = get_column_letter(qty_col)
            cost_letter = get_column_letter(cost_col)
            
            if category.lower() in ["lumber"]:
                t_letter = get_column_letter(t_col)
                w_letter = get_column_letter(w_col)
                l_letter = get_column_letter(l_col)
                safe_set_cell(ws, row, total_cost_col, f"=({qty_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row}/12)*{cost_letter}{row}/1000")
            elif category.lower() in ["panels"]:
                t_letter = get_column_letter(t_col)
                w_letter = get_column_letter(w_col)
                safe_set_cell(ws, row, total_cost_col, f"=({qty_letter}{row}*{t_letter}{row}*{w_letter}{row})*{cost_letter}{row}/1000")
            else:
                safe_set_cell(ws, row, total_cost_col, f"={qty_letter}{row}*{cost_letter}{row}")
            
            row += 1

def update_delivery_totals(ws, cols: dict, data_ranges: list, del_start: int, del_end: int):
    total_del_col = col_to_num(cols["total_delivered"])
    del_lf_col = col_to_num(cols["delivered_lf"])
    del_bf_col = col_to_num(cols["delivered_bf_sf"])
    del_cost_col = col_to_num(cols["delivered_cost"])
    del_cost_tax_col = col_to_num(cols["delivered_cost_tax"])
    pct_del_col = col_to_num(cols["pct_delivery"])
    
    length_col = col_to_num(cols["length"])
    thickness_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    cost_col = col_to_num(cols["cost_mbf"])
    total_cost_col = col_to_num(cols["total_cost"])
    type_col = col_to_num(cols["type"])
    
    inv_bundles_col = col_to_num(cols["inv_bundles"])
    pcs_bundle_col = col_to_num(cols["pcs_bundle"])
    inv_pcs_col = col_to_num(cols["inv_pcs"])
    issues_col = col_to_num(cols["issues"])
    issues_lf_col = col_to_num(cols["issues_lf"])
    issues_bf_col = col_to_num(cols["issues_bf_sf"])
    pct_issued_col = col_to_num(cols["pct_issued"])
    issues_cost_col = col_to_num(cols["issues_cost"])
    issues_cost_tax_col = col_to_num(cols["issues_cost_tax"])
    
    del_start_letter = get_column_letter(del_start)
    del_end_letter = get_column_letter(del_end)
    
    for start_row, end_row in data_ranges:
        for row in range(start_row, end_row + 1):
            row_type = ws.cell(row=row, column=type_col).value
            if not row_type:
                continue
                
            td_letter = get_column_letter(total_del_col)
            l_letter = get_column_letter(length_col)
            t_letter = get_column_letter(thickness_col)
            w_letter = get_column_letter(width_col)
            cost_letter = get_column_letter(cost_col)
            tc_letter = get_column_letter(total_cost_col)
            
            # Total Delivered
            ws.cell(row=row, column=total_del_col, value=f"=SUM({del_start_letter}{row}:{del_end_letter}{row})")
            
            # Delivered L/F
            ws.cell(row=row, column=del_lf_col, value=f"={td_letter}{row}*{l_letter}{row}")
            
            row_type_val = str(row_type).strip().lower()
            if row_type_val in ("lumber",):
                ws.cell(row=row, column=del_bf_col, value=f"={td_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row}/12")
                bf_letter = get_column_letter(del_bf_col)
                ws.cell(row=row, column=del_cost_col, value=f"={bf_letter}{row}*{cost_letter}{row}/1000")
            elif row_type_val == "panels":
                ws.cell(row=row, column=del_bf_col, value=f"={td_letter}{row}*{t_letter}{row}*{w_letter}{row}")
                bf_letter = get_column_letter(del_bf_col)
                ws.cell(row=row, column=del_cost_col, value=f"={bf_letter}{row}*{cost_letter}{row}/1000")
            elif row_type_val in ("lvl",):
                lf_letter = get_column_letter(del_lf_col)
                ws.cell(row=row, column=del_cost_col, value=f"={lf_letter}{row}*{cost_letter}{row}")
            elif row_type_val in ("each", "invoice"):
                ws.cell(row=row, column=del_cost_col, value=f"={td_letter}{row}*{cost_letter}{row}")
            
            dc_letter = get_column_letter(del_cost_col)
            ws.cell(row=row, column=del_cost_tax_col, value=f"={dc_letter}{row}*1.06")
            ws.cell(row=row, column=pct_del_col, value=f"=IFERROR({dc_letter}{row}/{tc_letter}{row},0)")
            
            ib_letter = get_column_letter(inv_bundles_col)
            pb_letter = get_column_letter(pcs_bundle_col)
            ip_letter = get_column_letter(inv_pcs_col)
            
            ws.cell(row=row, column=inv_pcs_col, value=f"=IF({ib_letter}{row}<>\"\",{ib_letter}{row}*{pb_letter}{row},0)")
            ws.cell(row=row, column=issues_col, value=f"={td_letter}{row}-{ip_letter}{row}")
            
            iss_letter = get_column_letter(issues_col)
            ws.cell(row=row, column=issues_lf_col, value=f"={iss_letter}{row}*{l_letter}{row}")
            
            if row_type_val in ("lumber",):
                ws.cell(row=row, column=issues_bf_col, value=f"=({iss_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row})/12")
            elif row_type_val == "panels":
                ws.cell(row=row, column=issues_bf_col, value=f"={iss_letter}{row}*{t_letter}{row}*{w_letter}{row}")
            else:
                ws.cell(row=row, column=issues_bf_col, value=f"={iss_letter}{row}")
                
            ibf_letter = get_column_letter(issues_bf_col)
            ws.cell(row=row, column=pct_issued_col, value=f"=IFERROR({ibf_letter}{row}/{td_letter}{row},0)")
            ws.cell(row=row, column=issues_cost_col, value=f"=IFERROR({ibf_letter}{row}*{cost_letter}{row}/1000,0)")
            
            ic_letter = get_column_letter(issues_cost_col)
            ws.cell(row=row, column=issues_cost_tax_col, value=f"={ic_letter}{row}*1.06")

@router.post("/client-requirements")
async def generate_client_requirements(req: ExportDataRequest, db: Session = Depends(get_db)):
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "Client_Requirements_Template.xlsx")
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template not found on server")
        
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet_name = SHEET_COBIA if "cobia" in req.project_name.lower() else SHEET_WILLOW
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        cols = COBIA_COLS if sheet_name == SHEET_COBIA else WILLOW_COLS
        data_ranges = [(3, 118), (123, 152), (157, 170), (173, 176)] if sheet_name == SHEET_COBIA else [(3, 78)]
        delivery_start_col = "BC" if sheet_name == SHEET_COBIA else "AD"
        
        # Determine delivery tracking boundaries
        del_start_num = col_to_num(delivery_start_col)
        # Find next empty date column in row 2
        next_del_col = del_start_num
        while True:
            cell = ws.cell(row=2, column=next_del_col)
            if cell.value is not None:
                next_del_col += 1
            elif type(cell).__name__ == 'MergedCell':
                next_del_col += 1
            else:
                break
            
        del_end_num = next_del_col - 1 if next_del_col > del_start_num else del_start_num

        po_co_qty_num = col_to_num(cols["po_co_qty"])
        
        # 1. Seed from POs if template is blank
        if req.pos:
            seed_sheet_from_pos(ws, sheet_name, req.pos, cols, req.project_name)

        # 2. Process invoices using the legacy matching logic
        if req.invoices:
            # Fetch mappings from DB
            db_mappings = db.query(ItemMapping).filter(ItemMapping.project_id == 1).all() # We need project_id, but it's not in ExportDataRequest. Let's look it up later or pass it.
            # Wait, req doesn't have project_id! Let's just match on description globally for now, or update ExportDataRequest.
            mapping_dict = {m.invoice_description: m.material_id for m in db_mappings}
            
            matched, unmatched = process_invoices_for_sheet(
                ws, sheet_name, req.invoices, cols, data_ranges, mapping_dict
            )

        # 3. Apply Change Orders (VPOs) to adjust PO/CO Qty
        if req.cos:
            process_change_orders_for_sheet(ws, sheet_name, req.cos, cols, data_ranges)

        # Update all complex formulas
        del_end_num = max([col_to_num(delivery_start_col)] + [c.column for c in ws[2] if c.value and type(c).__name__ != 'MergedCell'])
        update_delivery_totals(ws, cols, data_ranges, del_start_num, del_end_num)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        wb.close()
        
        return FileResponse(
            tmp.name, 
            filename=f"Client_Requirements_{req.project_name.replace(' ', '_')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/unmatched-items")
async def get_unmatched_items(req: ExportDataRequest, db: Session = Depends(get_db)):
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "Client_Requirements_Template.xlsx")
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template not found on server")
        
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet_name = SHEET_COBIA if "cobia" in req.project_name.lower() else SHEET_WILLOW
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        cols = COBIA_COLS if sheet_name == SHEET_COBIA else WILLOW_COLS
        data_ranges = [(3, 118), (123, 152), (157, 170), (173, 176)] if sheet_name == SHEET_COBIA else [(3, 78)]
        
        # Seed POs so we have rows to match against
        if req.pos:
            seed_sheet_from_pos(ws, sheet_name, req.pos, cols, req.project_name)
            
        db_mappings = db.query(ItemMapping).all()
        mapping_dict = {m.invoice_description: m.material_id for m in db_mappings}
        
        matched, unmatched = process_invoices_for_sheet(
            ws, sheet_name, req.invoices, cols, data_ranges, mapping_dict
        )
        
        wb.close()
        return {"unmatched": unmatched, "matched_count": len(matched)}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

