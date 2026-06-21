"""
Excel Importer Service — Full Implementation
Reads Client_Requirments_Doc.xlsx and populates the database with:
  - Material rows (type, qty, dims, cost_mbf, material_type)
  - CO adjustments (co_qty per material per CO column)
  - Delivery records (per delivery-date column)
  - Inventory data (bundles, pcs_per_bundle, inv_pcs, issues, variance_code, reason)
  - VPO entries from the VPO sheet
"""
import os
from datetime import datetime
from typing import List, Tuple, Optional, Dict

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import Project, Material, Inventory, Delivery, COAdjustment, VPO


# ─────────────────────────────────────────────────────────────────
# Column maps — exact mirrors of generate_requirements.py
# ─────────────────────────────────────────────────────────────────

WILLOW_COLS = {
    "type": "A", "qty": "B",
    "co_qty": "Q", "po_co_qty": "R",
    "thickness": "S", "width": "U", "length": "V",
    "material_type": "W", "lf_pcs": "X", "bf_sf": "Y",
    "cost_mbf": "Z", "total_cost": "AA", "total_cost_tax": "AB",
    "invoice_num": "AC",
    "delivery_start": "AD",
    "total_delivered": "AV", "delivered_lf": "AW",
    "delivered_bf_sf": "AX", "delivered_cost": "AY", "delivered_cost_tax": "AZ",
    "pct_delivery": "BA", "inv_bundles": "BB", "inv_uom": "BC",
    "pcs_bundle": "BD", "inv_pcs": "BE", "issues": "BF",
    "issues_lf": "BG", "issues_bf_sf": "BH", "pct_issued": "BI",
    "issues_cost": "BJ", "issues_cost_tax": "BK",
    "variance_code": "BL", "reason": "BM",
}

COBIA_COLS = {
    "type": "A", "qty": "B",
    "co_qty": "AP", "po_co_qty": "AQ",
    "thickness": "AR", "width": "AT", "length": "AU",
    "material_type": "AV", "lf_pcs": "AW", "bf_sf": "AX",
    "cost_mbf": "AY", "total_cost": "AZ", "total_cost_tax": "BA",
    "invoice_num": "BB",
    "delivery_start": "BC",
    "total_delivered": "DV", "delivered_lf": "DW",
    "delivered_bf_sf": "DX", "delivered_cost": "DY", "delivered_cost_tax": "DZ",
    "pct_delivery": "EA", "inv_bundles": "EB", "inv_uom": "EC",
    "pcs_bundle": "ED", "inv_pcs": "EE", "issues": "EF",
    "issues_lf": "EG", "issues_bf_sf": "EH", "pct_issued": "EI",
    "issues_cost": "EJ", "issues_cost_tax": "EK",
    "variance_code": "EL", "reason": "EM",
}

# CO header columns (between qty and co_qty_total)
WILLOW_CO_RANGE = ("C", "P")   # cols C–P are individual CO columns
COBIA_CO_RANGE  = ("C", "AO")  # cols C–AO

WILLOW_DATA_RANGES: List[Tuple[int, int]] = [(3, 78)]
COBIA_DATA_RANGES:  List[Tuple[int, int]] = [(3, 118), (123, 152), (157, 170), (173, 176)]


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def col_to_num(col: str) -> int:
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", "").replace("$", "")
        if s in ("", "—", "-", "N/A", "#REF!", "#VALUE!", "#DIV/0!"):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("-", "—", "None") else s


def parse_date(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if hasattr(v, 'year'):  # date object
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d-%b-%y"]:
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass
    return None


def get_sheet(wb, project_name: str):
    name_upper = project_name.upper()
    prefer = "Cobia Cove Appartments" if "COBIA" in name_upper else "Willow Way Apts"
    if prefer in wb.sheetnames:
        return wb[prefer], prefer
    for sn in wb.sheetnames:
        su = sn.upper()
        if "COBIA" in name_upper and "COBIA" in su:
            return wb[sn], sn
        if "WILLOW" in name_upper and "WILLOW" in su:
            return wb[sn], sn
    return None, None


def _find_delivery_cols(ws, delivery_start: str, total_delivered_col: str, date_row: int = 2) -> Dict[int, datetime]:
    """Return {col_num: ship_date} for every date-valued header cell between delivery_start and total_delivered."""
    start_n = col_to_num(delivery_start)
    end_n   = col_to_num(total_delivered_col) - 1
    result  = {}
    for cn in range(start_n, end_n + 1):
        val = ws.cell(row=date_row, column=cn).value
        dt  = parse_date(val)
        if dt:
            result[cn] = dt
    return result


def _find_co_cols(ws, co_start: str, co_end: str, header_row: int = 2) -> Dict[int, str]:
    """Return {col_num: co_label} for columns with non-empty header in the CO range."""
    start_n = col_to_num(co_start)
    end_n   = col_to_num(co_end)
    result  = {}
    for cn in range(start_n, end_n + 1):
        val = ws.cell(row=header_row, column=cn).value
        if val is not None:
            label = str(val).strip()
            if label:
                result[cn] = label
    return result


# ─────────────────────────────────────────────────────────────────
# Main importer
# ─────────────────────────────────────────────────────────────────

def import_excel_data(db: Session, project: Project, excel_path: str) -> dict:
    """
    Full import of an Excel file into the database.
    Clears existing data for the project first, then re-imports.
    Returns a summary dict with counts.
    """
    if not os.path.exists(excel_path):
        return {"error": f"File not found: {excel_path}"}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws, sheet_name = get_sheet(wb, project.name)
    if ws is None:
        return {"error": f"No matching sheet found. Available: {wb.sheetnames}"}

    is_cobia    = "COBIA" in project.name.upper()
    cols        = COBIA_COLS if is_cobia else WILLOW_COLS
    data_ranges = COBIA_DATA_RANGES if is_cobia else WILLOW_DATA_RANGES
    co_range    = COBIA_CO_RANGE if is_cobia else WILLOW_CO_RANGE

    # Discover dynamic columns from row 2 headers
    delivery_date_cols = _find_delivery_cols(ws, cols["delivery_start"], cols["total_delivered"], date_row=2)
    co_header_cols     = _find_co_cols(ws, co_range[0], co_range[1], header_row=2)

    # ── Wipe existing project data ────────────────────────────────
    for m in db.query(Material).filter(Material.project_id == project.id).all():
        db.query(Inventory).filter(Inventory.material_id == m.id).delete()
        db.query(Delivery).filter(Delivery.material_id == m.id).delete()
        db.query(COAdjustment).filter(COAdjustment.material_id == m.id).delete()
        db.delete(m)
    db.query(VPO).filter(VPO.project_id == project.id).delete()
    db.commit()

    mat_count = deliv_count = co_count = inv_count = vpo_count = 0

    for start_row, end_row in data_ranges:
        for row in range(start_row, end_row + 1):
            type_val = ws.cell(row=row, column=col_to_num(cols["type"])).value
            if not type_val or str(type_val).strip() == "":
                continue
            row_type = str(type_val).strip()

            # ── Core material fields ──────────────────────────────
            qty      = safe_float(ws.cell(row=row, column=col_to_num(cols["qty"])).value) or 0
            co_qty   = safe_float(ws.cell(row=row, column=col_to_num(cols["co_qty"])).value) or 0
            po_co    = safe_float(ws.cell(row=row, column=col_to_num(cols["po_co_qty"])).value) or (qty + co_qty)
            thick    = safe_float(ws.cell(row=row, column=col_to_num(cols["thickness"])).value)
            width    = safe_float(ws.cell(row=row, column=col_to_num(cols["width"])).value)
            length   = safe_float(ws.cell(row=row, column=col_to_num(cols["length"])).value)
            mat_type = safe_str(ws.cell(row=row, column=col_to_num(cols["material_type"])).value)
            lf_pcs   = safe_float(ws.cell(row=row, column=col_to_num(cols["lf_pcs"])).value) or 0
            bf_sf    = safe_float(ws.cell(row=row, column=col_to_num(cols["bf_sf"])).value) or 0
            cost_mbf = safe_float(ws.cell(row=row, column=col_to_num(cols["cost_mbf"])).value) or 0
            tot_cost = safe_float(ws.cell(row=row, column=col_to_num(cols["total_cost"])).value) or 0
            tot_tax  = safe_float(ws.cell(row=row, column=col_to_num(cols["total_cost_tax"])).value) or 0
            inv_refs = safe_str(ws.cell(row=row, column=col_to_num(cols["invoice_num"])).value)

            mat = Material(
                project_id=project.id,
                type=row_type,
                qty=qty, co_qty=co_qty, po_co_qty=po_co,
                thickness=thick, width=width, length=length,
                material_type=mat_type,
                lf_pcs=lf_pcs, bf_sf=bf_sf,
                cost_mbf=cost_mbf, total_cost=tot_cost, total_cost_tax=tot_tax,
                invoice_refs=inv_refs,
            )
            db.add(mat)
            db.flush()
            mat_count += 1

            # ── CO adjustments ────────────────────────────────────
            for cn, co_label in co_header_cols.items():
                val = safe_float(ws.cell(row=row, column=cn).value)
                if val is not None and val != 0:
                    db.add(COAdjustment(
                        material_id=mat.id,
                        co_number=co_label,
                        co_date="",
                        qty_change=val,
                        description=f"{co_label} — {mat_type}",
                    ))
                    co_count += 1

            # ── Deliveries (per date column) ──────────────────────
            for cn, ship_date in delivery_date_cols.items():
                qty_d = safe_float(ws.cell(row=row, column=cn).value)
                if qty_d and qty_d > 0:
                    db.add(Delivery(
                        material_id=mat.id,
                        document_id=None,
                        invoice_number=inv_refs,
                        ship_date=ship_date,
                        quantity=qty_d,
                        qty_multiplier=1.0,
                        uom="PC",
                    ))
                    deliv_count += 1

            # ── Inventory / Issues ────────────────────────────────
            bundles    = safe_float(ws.cell(row=row, column=col_to_num(cols["inv_bundles"])).value)
            inv_uom    = safe_str(ws.cell(row=row, column=col_to_num(cols["inv_uom"])).value)
            pcs_bundle = safe_float(ws.cell(row=row, column=col_to_num(cols["pcs_bundle"])).value)
            inv_pcs    = safe_float(ws.cell(row=row, column=col_to_num(cols["inv_pcs"])).value) or 0
            issues_val = safe_float(ws.cell(row=row, column=col_to_num(cols["issues"])).value) or 0
            var_code   = safe_str(ws.cell(row=row, column=col_to_num(cols["variance_code"])).value)
            reason_val = safe_str(ws.cell(row=row, column=col_to_num(cols["reason"])).value)

            if any([bundles, inv_pcs != 0, issues_val != 0, var_code, reason_val]):
                db.add(Inventory(
                    material_id=mat.id,
                    bundles=bundles, uom=inv_uom, pcs_per_bundle=pcs_bundle,
                    inv_pcs=inv_pcs, issues=issues_val,
                    variance_code=var_code, reason=reason_val,
                ))
                inv_count += 1

    db.commit()

    # ── VPO Sheet ─────────────────────────────────────────────────
    for vpo_sheet in ["VPO's", "VPOs", "VPO"]:
        if vpo_sheet in wb.sheetnames:
            vws = wb[vpo_sheet]
            for vrow in range(3, vws.max_row + 1):
                desc = safe_str(vws.cell(row=vrow, column=2).value)
                if not desc:
                    continue
                db.add(VPO(
                    project_id=project.id,
                    vpo_date=parse_date(vws.cell(row=vrow, column=1).value),
                    qty=safe_float(vws.cell(row=vrow, column=3).value) or 0,
                    uom=safe_str(vws.cell(row=vrow, column=4).value),
                    description=desc,
                    footage=safe_float(vws.cell(row=vrow, column=5).value) or 0,
                    price=safe_float(vws.cell(row=vrow, column=6).value) or 0,
                    amount=safe_float(vws.cell(row=vrow, column=7).value) or 0,
                    tax=safe_float(vws.cell(row=vrow, column=8).value) or 0,
                    total=safe_float(vws.cell(row=vrow, column=9).value) or 0,
                    co_ref=safe_str(vws.cell(row=vrow, column=10).value),
                    co_number=safe_str(vws.cell(row=vrow, column=11).value),
                    remarks=safe_str(vws.cell(row=vrow, column=12).value),
                ))
                vpo_count += 1
            db.commit()
            break

    return {
        "sheet": sheet_name,
        "materials": mat_count,
        "co_adjustments": co_count,
        "deliveries": deliv_count,
        "inventory_records": inv_count,
        "vpos": vpo_count,
    }
