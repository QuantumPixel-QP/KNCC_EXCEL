import openpyxl
import re
from openpyxl.utils import get_column_letter

def col_to_num(col_str: str) -> int:
    return openpyxl.utils.cell.column_index_from_string(col_str)

def classify_item_category(desc: str, code: str = "") -> str:
    desc = desc.upper()
    code = str(code).upper()
    if "OSB" in desc or "PLYWOOD" in desc or "ZIP" in desc or "ADVANTECH" in desc:
        return "panels"
    elif "GLULAM" in desc or "LVL" in desc or "EWP" in desc or "MICROLLAM" in desc or "PARALLAM" in desc:
        return "lvl"
    elif "HUC" in desc or "SIMPSON" in desc or "LSTA" in desc or "H2.5A" in desc:
        return "hardware"
    elif "TREX" in desc or "SEAL" in desc or "TAPE" in desc or "WRAP" in desc or "ADHESIVE" in desc:
        return "each"
    return "lumber"

def safe_set_cell(ws, row, col, val):
    try:
        ws.cell(row=row, column=col, value=val)
    except AttributeError:
        pass # Ignore MergedCell error

def extract_date_columns(ws, start_col: int, date_row: int = 2) -> dict:
    date_cols = {}
    for col in range(start_col, ws.max_column + 1):
        try:
            val = ws.cell(row=date_row, column=col).value
            if val:
                from datetime import datetime
                if isinstance(val, datetime):
                    ds = val.strftime("%m/%d/%Y")
                    date_cols[ds] = col
                elif isinstance(val, str) and "/" in val:
                    try:
                        dt = datetime.strptime(val.strip(), "%m/%d/%y")
                        date_cols[dt.strftime("%m/%d/%Y")] = col
                    except ValueError:
                        try:
                            dt = datetime.strptime(val.strip(), "%m/%d/%Y")
                            date_cols[dt.strftime("%m/%d/%Y")] = col
                        except ValueError:
                            date_cols[val.strip()] = col
        except AttributeError:
            pass # MergedCell
    return date_cols

def find_or_add_delivery_date_col(ws, ship_date: str, date_cols: dict, date_row: int = 2):
    if not ship_date:
        return list(date_cols.values())[-1] if date_cols else None
        
    from datetime import datetime
    try:
        dt = datetime.strptime(ship_date, "%Y-%m-%d")
        ship_date_norm = dt.strftime("%m/%d/%Y")
    except ValueError:
        ship_date_norm = ship_date

    if ship_date_norm in date_cols:
        return date_cols[ship_date_norm]
        
    if not date_cols:
        return None

    try:
        dt_ship = datetime.strptime(ship_date_norm, "%m/%d/%Y")
        sorted_dates = sorted(
            [d for d in date_cols.keys()], 
            key=lambda x: datetime.strptime(x, "%m/%d/%Y") if "/" in x else datetime.min
        )
        insert_after = None
        for d in sorted_dates:
            try:
                if datetime.strptime(d, "%m/%d/%Y") < dt_ship:
                    insert_after = d
            except:
                pass
                
        max_date_col = max(date_cols.values())
        new_col = max_date_col + 1
        
        while type(ws.cell(row=date_row, column=new_col)).__name__ == 'MergedCell':
            new_col += 1
            
        ws.cell(row=date_row, column=new_col, value=ship_date_norm)
        date_cols[ship_date_norm] = new_col
        return new_col
    except Exception:
        max_date_col = max(date_cols.values())
        new_col = max_date_col + 1
        while type(ws.cell(row=date_row, column=new_col)).__name__ == 'MergedCell':
            new_col += 1
        ws.cell(row=date_row, column=new_col, value=ship_date_norm)
        date_cols[ship_date_norm] = new_col
        return new_col

def update_invoice_reference(ws, row: int, col_num: int, invoice_number: str):
    cell = ws.cell(row=row, column=col_num)
    current = cell.value
    if not current:
        cell.value = str(invoice_number)
    else:
        curr_str = str(current)
        if str(invoice_number) not in curr_str:
            cell.value = f"{curr_str}, {invoice_number}"

def process_invoices_for_sheet(ws, sheet_name: str, invoices: list, cols: dict, data_ranges: list, mappings: dict = None):
    if mappings is None:
        mappings = {}
        
    date_start_num = col_to_num(cols.get("delivery_start", "BC" if "cobia" in sheet_name.lower() else "AD"))
    date_cols = extract_date_columns(ws, date_start_num, 2)
    inv_col = col_to_num(cols["invoices"])
    
    t_col = col_to_num(cols["thickness"])
    w_col = col_to_num(cols["width"])
    l_col = col_to_num(cols["length"])
    type_col = col_to_num(cols["type"])
    desc_col = col_to_num(cols["material_type"])
    
    matched_items = []
    unmatched_items = []
    
    for inv in invoices:
        ship_date = inv.get("date", "2026-06-25")
        invoice_number = inv.get("invoice_number", "Unknown")
        
        for item in inv.get("line_items", []):
            inv_desc = item.get("description", "").upper()
            
            # Check manual mappings first
            if inv_desc in mappings:
                target_row = mappings[inv_desc]
                
                date_col = find_or_add_delivery_date_col(ws, ship_date, date_cols, 2)
                if not date_col:
                    date_col = date_start_num
                    
                curr_qty = ws.cell(row=target_row, column=date_col).value or 0
                qty = float(item.get('quantity', 0))
                safe_set_cell(ws, target_row, date_col, curr_qty + qty)
                
                update_invoice_reference(ws, target_row, inv_col, invoice_number)
                matched_items.append(item)
                continue
            
            inv_dims = item.get("dimensions", "").strip()
            inv_t, inv_w, inv_l = None, None, None
            
            dm = re.match(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_dims)
            if not dm:
                dm = re.search(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_desc)
                
            if dm:
                inv_t = int(dm.group(1))
                inv_w = int(dm.group(2))
                inv_l = int(dm.group(3))
            else:
                dm_2d = re.search(r'(\d+)\s*[xX]\s*(\d+)', inv_desc)
                if dm_2d:
                    inv_w = int(dm_2d.group(1))
                    inv_l = int(dm_2d.group(2))
                    
            if not inv_l and "PET 104-5/8" in inv_desc:
                inv_l = 9
            elif not inv_l and "PET 116-5/8" in inv_desc:
                inv_l = 10
            
            inv_category = classify_item_category(inv_desc, item.get("item_code", ""))
            
            best_match = None
            best_score = 0
            
            for start_row, end_row in data_ranges:
                for r in range(start_row, end_row + 1):
                    row_type = ws.cell(row=r, column=type_col).value
                    if not row_type:
                        continue
                        
                    rt_norm = str(row_type).strip().lower()
                    if rt_norm != inv_category and not (rt_norm == "panels" and inv_category == "lumber"):
                        continue
                        
                    row_t = ws.cell(row=r, column=t_col).value
                    row_w = ws.cell(row=r, column=w_col).value
                    row_l = ws.cell(row=r, column=l_col).value
                    row_mat = str(ws.cell(row=r, column=desc_col).value or "").upper()
                    
                    score = 0
                    if rt_norm == inv_category: score += 10
                    else: score += 5
                    
                    if inv_t is not None and row_t is not None and inv_t == row_t: score += 5
                    if inv_w is not None and row_w is not None and inv_w == row_w: score += 5
                    if inv_l is not None and row_l is not None and inv_l == row_l: score += 5
                    
                    if "PT" in inv_desc and ("PT" in row_mat or "MCA" in row_mat): score += 3
                    elif "MCA" in inv_desc and ("MCA" in row_mat or "PT" in row_mat): score += 3
                    elif "SYP" in inv_desc and "SYP" in row_mat: score += 3
                    elif "LVL" in inv_desc and "LVL" in row_mat: score += 3
                    elif "OSB" in inv_desc and "OSB" in row_mat: score += 3
                    elif "TREX" in inv_desc and "TREX" in row_mat: score += 3
                    
                    if score > best_score:
                        best_score = score
                        best_match = r
                        
            if best_score >= 15 and best_match:
                date_col = find_or_add_delivery_date_col(ws, ship_date, date_cols, 2)
                if not date_col:
                    date_col = date_start_num
                    
                curr_qty = ws.cell(row=best_match, column=date_col).value or 0
                qty = float(item.get('quantity', 0))
                safe_set_cell(ws, best_match, date_col, curr_qty + qty)
                
                update_invoice_reference(ws, best_match, inv_col, invoice_number)
                matched_items.append(item)
            else:
                unmatched_items.append({
                    "invoice_number": invoice_number,
                    "date": ship_date,
                    **item
                })
                
    return matched_items, unmatched_items
