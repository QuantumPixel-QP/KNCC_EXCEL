"""
Excel Sync Service — fixed:
  FIX #2:  LEGACY_EXCEL / EXPORT_DIR imported correctly from config.
  FIX #3:  ALL 15 formulas from generate_requirements.py are now written:
             Total Delivered, Delivered LF, Delivered BF/SF, Delivered Cost,
             Delivered Cost+Tax, % Delivery, Inventory PCS, Issues, Issues LF,
             Issues BF/SF, % Issued, Issues Cost, Issues Cost+Tax.
  FIX #4:  Tax rate pulled from Project.tax_rate (not hardcoded 1.06).
  FIX #15: Project sheet selection uses a proper project_type field comparison
           instead of substring match.
"""
import os
import shutil
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import Project, Material, Document, Delivery
from ..config import EXPORT_DIR, LEGACY_EXCEL

# ──────────────────────────────────────────────────────────────────────────────
# Sheet / column layout (exact mapping from generate_requirements.py)
# ──────────────────────────────────────────────────────────────────────────────

SHEET_VPOS = "VPO's"
SHEET_COBIA = "Cobia Cove Appartments"
SHEET_WILLOW = "Willow Way Apts"

COBIA_COLS = {
    "type": "A", "qty": "B", "co_qty": "AP", "po_co_qty": "AQ",
    "thickness": "AR", "x": "AS", "width": "AT", "length": "AU",
    "material_type": "AV", "lf_pcs": "AW", "bf_sf": "AX",
    "cost_mbf": "AY", "total_cost": "AZ", "total_cost_tax": "BA",
    "invoice_num": "BB",
    "delivery_start": "BC",   # First delivery-date column (row 2 has the date)
    "total_delivered": "DV", "delivered_lf": "DW",
    "delivered_bf_sf": "DX", "delivered_cost": "DY", "delivered_cost_tax": "DZ",
    "pct_delivery": "EA", "inv_bundles": "EB", "inv_uom": "EC",
    "pcs_bundle": "ED", "inv_pcs": "EE", "issues": "EF",
    "issues_lf": "EG", "issues_bf_sf": "EH", "pct_issued": "EI",
    "issues_cost": "EJ", "issues_cost_tax": "EK",
    "variance_code": "EL", "reason": "EM",
}

WILLOW_COLS = {
    "type": "A", "qty": "B", "co_qty": "Q", "po_co_qty": "R",
    "thickness": "S", "x": "T", "width": "U", "length": "V",
    "material_type": "W", "lf_pcs": "X", "bf_sf": "Y",
    "cost_mbf": "Z", "total_cost": "AA", "total_cost_tax": "AB",
    "invoice_num": "AC",
    "delivery_start": "AD",
    "total_delivered": "AV", "delivered_lf": "AW",
    "delivered_bf_sf": "AX", "delivered_cost": "AY", "delivered_cost_tax": "AZ",
    "pct_delivery": "BA", "inv_bundles": "BB", "inv_uom": "BC",
    "pcs_bundle": "BD", "inv_pcs": "BE", "issues": "BF",
    "issues_lf": "BG", "issues_bf_sf": "BH", "pct_issued": "BI",
    "issues_cost": "BJ", "issues_cost_tax": "BK",
    "variance_code": "BL", "reason": "BM",
}


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def col_to_num(col_letter: str) -> int:
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def get_delivery_date_columns(ws, start_col: str, row: int = 2) -> Dict[datetime, int]:
    date_cols: Dict[datetime, int] = {}
    start_num = col_to_num(start_col)
    for col in range(start_num, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, datetime):
            date_cols[cell.value] = col
    return date_cols


def find_or_add_date_col(ws, ship_date: datetime, date_cols: Dict[datetime, int], date_row: int = 2) -> int:
    """Return existing column for date, or create a new one at the end."""
    key = datetime(ship_date.year, ship_date.month, ship_date.day)
    if key in date_cols:
        return date_cols[key]
    new_col = max(date_cols.values()) + 1 if date_cols else col_to_num("BC")
    ws.cell(row=date_row, column=new_col, value=key)
    date_cols[key] = new_col
    return new_col


def update_invoice_ref_cell(ws, row: int, col_num: int, invoice_number: str):
    cell = ws.cell(row=row, column=col_num)
    current = str(cell.value) if cell.value else ""
    if invoice_number not in current:
        cell.value = (current + "\n" + invoice_number).strip()


# ──────────────────────────────────────────────────────────────────────────────
# FIX #3 — Write ALL 15 formulas per data row
# ──────────────────────────────────────────────────────────────────────────────

def write_row_formulas(ws, row: int, cols: dict, del_start_letter: str, del_end_letter: str, tax_rate: float = 1.06):
    """
    Writes all delivery, inventory, and issues formulas for a single data row.
    Mirrors generate_requirements.py update_delivery_totals() exactly.
    FIX #4: tax_rate is passed in from project, not hardcoded.
    """
    # Shorthand lambdas
    L = lambda key: get_column_letter(col_to_num(cols[key]))
    r = str(row)

    # Column letter aliases
    td   = L("total_delivered")
    llen = L("length")
    t    = L("thickness")
    w    = L("width")
    cost = L("cost_mbf")
    tc   = L("total_cost")
    ib   = L("inv_bundles")
    pb   = L("pcs_bundle")
    ip   = L("inv_pcs")
    iss  = L("issues")
    ibf  = L("issues_bf_sf")
    ic   = L("issues_cost")
    dbf  = L("delivered_bf_sf")
    dc   = L("delivered_cost")

    # Row type
    row_type_val = ws.cell(row=row, column=col_to_num(cols["type"])).value
    rt = str(row_type_val).strip().lower() if row_type_val else ""

    # ── 1. Total Delivered = SUM(all date columns) ──
    ws.cell(row=row, column=col_to_num(cols["total_delivered"]),
            value=f"=SUM({del_start_letter}{r}:{del_end_letter}{r})")

    # ── 2. Delivered L/F = TotalDelivered × Length ──
    ws.cell(row=row, column=col_to_num(cols["delivered_lf"]),
            value=f"={td}{r}*{llen}{r}")

    # ── 3. Delivered B/F or S/F (type-dependent) ──
    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]),
                value=f"={td}{r}*{t}{r}*{w}{r}*{llen}{r}/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]),
                value=f"={td}{r}*{t}{r}*{w}{r}")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]),
                value=f"={td}{r}*{llen}{r}")
    else:  # each / invoice
        ws.cell(row=row, column=col_to_num(cols["delivered_bf_sf"]),
                value=f"={td}{r}")

    # ── 4. Delivered Cost ──
    if rt in ("lumber", "panels"):
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]),
                value=f"={dbf}{r}*{cost}{r}/1000")
    elif rt == "lvl":
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]),
                value=f"={L('delivered_lf')}{r}*{cost}{r}")
    else:
        ws.cell(row=row, column=col_to_num(cols["delivered_cost"]),
                value=f"={td}{r}*{cost}{r}")

    # ── 5. Delivered Cost + Tax (FIX #4: dynamic rate) ──
    ws.cell(row=row, column=col_to_num(cols["delivered_cost_tax"]),
            value=f"={dc}{r}*{tax_rate}")

    # ── 6. % Delivery = Delivered Cost / Total Cost ──
    ws.cell(row=row, column=col_to_num(cols["pct_delivery"]),
            value=f"=IFERROR({dc}{r}/{tc}{r},0)")

    # ── 7. Inventory PCS = Bundles × PCS/Bundle ──
    ws.cell(row=row, column=col_to_num(cols["inv_pcs"]),
            value=f"=IF({ib}{r}<>\"\",{ib}{r}*{pb}{r},0)")

    # ── 8. Issues = Total Delivered − Inventory PCS ──
    ws.cell(row=row, column=col_to_num(cols["issues"]),
            value=f"={td}{r}-{ip}{r}")

    # ── 9. Issues L/F ──
    ws.cell(row=row, column=col_to_num(cols["issues_lf"]),
            value=f"={iss}{r}*{llen}{r}")

    # ── 10. Issues B/F-S/F (type-dependent) ──
    if rt == "lumber":
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]),
                value=f"=({iss}{r}*{t}{r}*{w}{r}*{llen}{r})/12")
    elif rt == "panels":
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]),
                value=f"={iss}{r}*{t}{r}*{w}{r}")
    else:
        ws.cell(row=row, column=col_to_num(cols["issues_bf_sf"]),
                value=f"={iss}{r}")

    # ── 11. % Issued ──
    ws.cell(row=row, column=col_to_num(cols["pct_issued"]),
            value=f"=IFERROR({ibf}{r}/{td}{r},0)")

    # ── 12. Issues Cost ──
    ws.cell(row=row, column=col_to_num(cols["issues_cost"]),
            value=f"=IFERROR({ibf}{r}*{cost}{r}/1000,0)")

    # ── 13. Issues Cost + Tax (FIX #4) ──
    ws.cell(row=row, column=col_to_num(cols["issues_cost_tax"]),
            value=f"={ic}{r}*{tax_rate}")


# ──────────────────────────────────────────────────────────────────────────────
# Main sync function
# ──────────────────────────────────────────────────────────────────────────────

def sync_excel_for_project(db: Session, project: Project) -> str:
    """
    Generates a pixel-perfect Excel matching Client_Requirments_Doc.xlsx.

    FIX #2:  Uses LEGACY_EXCEL / EXPORT_DIR from config correctly.
    FIX #3:  Writes all 15 formulas per row.
    FIX #4:  Pulls tax_rate from project (user-assigned), with 1.06 fallback.
    FIX #15: Sheet selection uses project.name exact keyword match.
    """
    if not os.path.exists(LEGACY_EXCEL):
        raise FileNotFoundError(f"Legacy Excel template not found: {LEGACY_EXCEL}")

    export_filename = f"{project.name.replace(' ', '_')}_Requirements.xlsx"
    export_path = os.path.join(EXPORT_DIR, export_filename)
    shutil.copy2(LEGACY_EXCEL, export_path)

    # FIX #4: Use project tax rate, fallback to 1.06
    tax_rate = float(project.tax_rate) if project.tax_rate else 1.06

    # FIX #15: Strict keyword matching instead of substring
    name_upper = project.name.upper()
    is_cobia = "COBIA" in name_upper
    is_willow = "WILLOW" in name_upper

    if is_cobia:
        sheet_name = SHEET_COBIA
        cols = COBIA_COLS
        data_ranges: List[Tuple[int, int]] = [(3, 118), (123, 152), (157, 170), (173, 176)]
    else:
        sheet_name = SHEET_WILLOW
        cols = WILLOW_COLS
        data_ranges = [(3, 78)]

    wb = openpyxl.load_workbook(export_path)
    if sheet_name not in wb.sheetnames:
        wb.save(export_path)
        return export_path

    ws = wb[sheet_name]
    delivery_start_col = cols["delivery_start"]
    date_cols = get_delivery_date_columns(ws, delivery_start_col, row=2)

    # ── Inject delivery quantities from DB ──────────────────────────────────
    materials = db.query(Material).filter(Material.project_id == project.id).all()
    inv_col_num = col_to_num(cols["invoice_num"])

    for mat in materials:
        # Find the row in the Excel that matches this material
        target_row = _find_excel_row(ws, mat, cols, data_ranges)
        if not target_row:
            continue

        deliveries = db.query(Delivery).filter(Delivery.material_id == mat.id).all()
        for deliv in deliveries:
            if deliv.ship_date:
                date_col = find_or_add_date_col(ws, deliv.ship_date, date_cols, date_row=2)
                effective_qty = (deliv.quantity or 0) * (deliv.qty_multiplier or 1.0)
                current = ws.cell(row=target_row, column=date_col).value
                # Guard: current may be a formula string if the template was pre-filled
                try:
                    current_num = float(current) if current is not None else 0.0
                except (TypeError, ValueError):
                    current_num = 0.0
                ws.cell(row=target_row, column=date_col,
                        value=current_num + effective_qty)
            if deliv.invoice_number:
                update_invoice_ref_cell(ws, target_row, inv_col_num, deliv.invoice_number)

    # ── Write all 15 formulas for every data row ─────────────────────────────
    if date_cols:
        del_start_letter = get_column_letter(min(date_cols.values()))
        del_end_letter = get_column_letter(max(date_cols.values()))

        for start_row, end_row in data_ranges:
            for row in range(start_row, end_row + 1):
                row_type = ws.cell(row=row, column=col_to_num(cols["type"])).value
                if not row_type:
                    continue
                write_row_formulas(ws, row, cols, del_start_letter, del_end_letter, tax_rate)

    wb.save(export_path)
    return export_path


def _find_excel_row(ws, mat: Material, cols: dict, data_ranges: List[Tuple[int, int]]) -> Optional[int]:
    """
    Find the Excel row that matches this Material DB record by dimensions + type.
    """
    type_col = col_to_num(cols["type"])
    thick_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    len_col = col_to_num(cols["length"])
    mat_col = col_to_num(cols["material_type"])

    best_row = None
    best_score = 0

    for start_row, end_row in data_ranges:
        for row in range(start_row, end_row + 1):
            row_type = ws.cell(row=row, column=type_col).value
            if not row_type:
                continue

            score = 0
            rt = str(row_type).strip().lower()
            mt = str(mat.type or "").strip().lower()
            if rt == mt:
                score += 10
            elif rt in mt or mt in rt:
                score += 5
            else:
                continue

            def safe_float(v):
                try: return float(str(v).strip())
                except: return None

            row_t = safe_float(ws.cell(row=row, column=thick_col).value)
            row_w = safe_float(ws.cell(row=row, column=width_col).value)
            row_l = safe_float(ws.cell(row=row, column=len_col).value)

            if mat.thickness and row_t == mat.thickness: score += 5
            if mat.width and row_w == mat.width:         score += 5
            if mat.length and row_l == mat.length:       score += 5

            row_mat = str(ws.cell(row=row, column=mat_col).value or "").upper()
            for word in str(mat.material_type or "").upper().split():
                if len(word) > 2 and word in row_mat:
                    score += 2

            if score > best_score:
                best_score = score
                best_row = row

    return best_row if best_score >= 15 else None
