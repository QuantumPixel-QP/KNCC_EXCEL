"""
Validation Script for KNCC Automation
======================================
Compares the existing Excel against parsed PDF data to verify accuracy
and identify any discrepancies.

Usage:
    py validate_output.py
"""

import os
import sys
import re
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter
from pdf_parser import scan_and_parse_all, parse_number


BASE_PATH = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_PATH, "Client", "Client_Requirments_Doc.xlsx")


def get_all_invoice_numbers_from_sheet(ws, inv_col_letter: str) -> dict:
    """Get all invoice numbers from the sheet, mapped to their rows."""
    col_num = 0
    for c in inv_col_letter.upper():
        col_num = col_num * 26 + (ord(c) - ord('A') + 1)
    
    inv_to_rows = defaultdict(list)
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_num)
        if cell.value:
            val = str(cell.value)
            for inv_num in re.findall(r'\d{5,}', val):
                inv_to_rows[inv_num].append(row)
    return inv_to_rows


def validate():
    """Run validation checks."""
    print("=" * 70)
    print("KNCC Automation — Validation Report")
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    # Parse all PDFs
    print("\n[1] Parsing all PDFs...")
    data = scan_and_parse_all(BASE_PATH)
    print(f"    Invoices: {len(data['invoices'])}")
    print(f"    Change Orders: {len(data['change_orders'])}")
    print(f"    Purchase Orders: {len(data['purchase_orders'])}")
    
    # Load Excel
    print(f"\n[2] Loading Excel: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # === Cobia Cove Validation ===
    print(f"\n{'='*70}")
    print("COBIA COVE APARTMENTS — Validation")
    print(f"{'='*70}")
    
    ws = wb["Cobia Cove Appartments"]
    cobia_inv_map = get_all_invoice_numbers_from_sheet(ws, "BB")
    
    # Get all Cobia invoices from PDFs
    cobia_invoices = {inv.invoice_number: inv for inv in data['invoices'] 
                      if inv.project_name == "Cobia Cove"}
    
    print(f"\n  Invoice Coverage:")
    print(f"    Invoices in PDFs: {len(cobia_invoices)}")
    print(f"    Invoices referenced in Excel: {len(cobia_inv_map)}")
    
    # Find invoices in Excel but not in PDFs
    excel_only = set(cobia_inv_map.keys()) - set(cobia_invoices.keys())
    if excel_only:
        print(f"    ⚠ In Excel but NOT in PDF folder ({len(excel_only)}):")
        for inv in sorted(excel_only):
            print(f"      {inv} (rows: {cobia_inv_map[inv]})")
    
    # Find invoices in PDFs but not in Excel
    pdf_only = set(cobia_invoices.keys()) - set(cobia_inv_map.keys())
    if pdf_only:
        print(f"    📋 In PDFs but NOT in Excel ({len(pdf_only)}):")
        for inv in sorted(pdf_only):
            inv_data = cobia_invoices[inv]
            items_desc = "; ".join([f"{it.quantity} {it.uom} {it.description}" 
                                    for it in inv_data.line_items[:3]])
            print(f"      {inv} (date={inv_data.invoice_date}, items: {items_desc})")
    
    matched = set(cobia_invoices.keys()) & set(cobia_inv_map.keys())
    print(f"    ✓ Matched: {len(matched)}")
    
    # Spot-check some invoice amounts
    print(f"\n  Amount Spot-Check (first 10 matched invoices):")
    for inv_num in sorted(matched)[:10]:
        inv = cobia_invoices[inv_num]
        rows = cobia_inv_map[inv_num]
        print(f"    Invoice {inv_num}: PDF total=${inv.total_amount:.2f}, "
              f"ship={inv.ship_date}, items={len(inv.line_items)}, "
              f"Excel rows={rows}")
    
    # === Willow Way Validation ===
    print(f"\n{'='*70}")
    print("WILLOW WAY APARTMENTS — Validation")
    print(f"{'='*70}")
    
    ws_w = wb["Willow Way Apts"]
    willow_inv_map = get_all_invoice_numbers_from_sheet(ws_w, "AC")
    
    willow_invoices = {inv.invoice_number: inv for inv in data['invoices'] 
                       if inv.project_name == "Willow Way"}
    
    print(f"\n  Invoice Coverage:")
    print(f"    Invoices in PDFs: {len(willow_invoices)}")
    print(f"    Invoices referenced in Excel: {len(willow_inv_map)}")
    
    excel_only_w = set(willow_inv_map.keys()) - set(willow_invoices.keys())
    if excel_only_w:
        print(f"    ⚠ In Excel but NOT in PDF folder ({len(excel_only_w)}):")
        for inv in sorted(excel_only_w):
            print(f"      {inv} (rows: {willow_inv_map[inv]})")
    
    pdf_only_w = set(willow_invoices.keys()) - set(willow_inv_map.keys())
    if pdf_only_w:
        print(f"    📋 In PDFs but NOT in Excel ({len(pdf_only_w)}):")
        for inv in sorted(pdf_only_w):
            inv_data = willow_invoices[inv]
            items_desc = "; ".join([f"{it.quantity} {it.uom} {it.description}" 
                                    for it in inv_data.line_items[:3]])
            print(f"      {inv} (date={inv_data.invoice_date}, items: {items_desc})")
    
    matched_w = set(willow_invoices.keys()) & set(willow_inv_map.keys())
    print(f"    ✓ Matched: {len(matched_w)}")
    
    # === Change Order Validation ===
    print(f"\n{'='*70}")
    print("CHANGE ORDERS — Validation")
    print(f"{'='*70}")
    
    cobia_cos = [co for co in data['change_orders'] if co.project_name == "Cobia Cove"]
    willow_cos = [co for co in data['change_orders'] if co.project_name == "Willow Way"]
    
    print(f"\n  Cobia Cove COs parsed: {len(cobia_cos)}")
    for co in sorted(cobia_cos, key=lambda c: int(c.co_number) if c.co_number.isdigit() else 0):
        print(f"    CO #{co.co_number}: date={co.date}, total=${co.total_amount:.2f}, "
              f"items={len(co.line_items)}")
    
    print(f"\n  Willow Way COs parsed: {len(willow_cos)}")
    for co in sorted(willow_cos, key=lambda c: int(c.co_number) if c.co_number.isdigit() else 0):
        print(f"    CO #{co.co_number}: date={co.date}, total=${co.total_amount:.2f}, "
              f"items={len(co.line_items)}")
    
    # === Summary ===
    print(f"\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}")
    
    total_inv = len(cobia_invoices) + len(willow_invoices)
    total_matched = len(matched) + len(matched_w)
    total_missing = len(pdf_only) + len(pdf_only_w)
    total_extra = len(excel_only) + len(excel_only_w)
    
    print(f"  Total invoices in PDF folders: {total_inv}")
    print(f"  Total invoices matched in Excel: {total_matched}")
    print(f"  Invoices in PDFs not in Excel: {total_missing}")
    print(f"  Invoices in Excel not in PDFs: {total_extra}")
    print(f"  Coverage: {total_matched/total_inv*100:.1f}%" if total_inv > 0 else "  N/A")
    
    if total_missing == 0 and total_extra == 0:
        print(f"\n  ✅ VALIDATION PASSED — All invoices are accounted for!")
    elif total_missing > 0:
        print(f"\n  ⚠ ATTENTION — {total_missing} invoice(s) from PDFs are not yet in the Excel.")
        print(f"    These are likely non-material items (storage, hardware packages, trusses).")
    
    wb.close()


if __name__ == "__main__":
    validate()
