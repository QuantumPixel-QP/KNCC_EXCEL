import openpyxl
wb = openpyxl.load_workbook('Client/Client_Requirments_Doc.xlsx')

# Get the Cobia Cove sheet details
ws = wb['Cobia Cove Appartments']
print(f"Sheet: Cobia Cove Appartments (rows={ws.max_row}, cols={ws.max_column})")

# Print ALL rows with data
for row_num in range(1, ws.max_row + 1):
    cells = []
    for cell in ws[row_num]:
        if cell.value is not None:
            cells.append(f"{cell.column_letter}:{repr(cell.value)}")
    if cells:
        print(f"  Row {row_num}: {' | '.join(cells[:15])}")  # First 15 columns only
