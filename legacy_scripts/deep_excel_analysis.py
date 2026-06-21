"""
Deep analysis of ALL sheets in the Client_Requirments_Doc.xlsx
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

# Use the BACKUP to see the original state
wb = openpyxl.load_workbook('Client/Client_Requirments_Doc_BACKUP.xlsx')

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: '{sheet_name}'")
    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
    print(f"  Merged cells: {len(list(ws.merged_cells.ranges))}")
    if ws.merged_cells.ranges:
        for mc in list(ws.merged_cells.ranges)[:10]:
            print(f"    {mc}")
    
    print(f"\n  --- ALL ROWS WITH DATA ---")
    for row_num in range(1, ws.max_row + 1):
        cells = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            if cell.value is not None:
                col_letter = get_column_letter(col)
                cells.append(f"{col_letter}:{repr(cell.value)[:60]}")
        if cells:
            # Limit output per row
            output = " | ".join(cells[:20])
            if len(cells) > 20:
                output += f" | ... ({len(cells)} total cells)"
            print(f"  Row {row_num}: {output}")

wb.close()
