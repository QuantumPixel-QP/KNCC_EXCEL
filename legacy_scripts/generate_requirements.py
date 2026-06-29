"""
KNCC Material Procurement Automation System
============================================
Reads the existing Client_Requirments_Doc.xlsx, scans all PDF invoices and change orders,
identifies new/unprocessed documents, and updates the Excel with:
- New delivery entries (from invoice ship dates + quantities)
- Updated delivery totals and cost calculations
- Automated inventory tracking (Delivered - Issued)
- All formulas preserved and extended

Usage:
    py generate_requirements.py [--dry-run]
"""

import os
import sys
import re
import copy
from datetime import datetime
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Set

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter
from pdf_parser import (
    scan_and_parse_all, InvoiceData, ChangeOrderData, LineItem,
    parse_number
)


# =============================================================================
#  CONFIGURATION
# =============================================================================

BASE_PATH = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_EXCEL = os.path.join(BASE_PATH, "Client", "Client_Requirments_Doc.xlsx")
OUTPUT_EXCEL = os.path.join(BASE_PATH, "Client", "Client_Requirments_Doc.xlsx")  # Overwrite in-place
BACKUP_EXCEL = os.path.join(BASE_PATH, "Client", "Client_Requirments_Doc_BACKUP.xlsx")

# Sheet names
SHEET_VPOS = "VPO's"
SHEET_COBIA = "Cobia Cove Appartments"
SHEET_WILLOW = "Willow Way Apts"

# Project to job number mapping
PROJECT_JOB_MAP = {
    "Cobia Cove": "68981",
    "Willow Way": "60126",
}

# Column layout for Cobia Cove sheet
COBIA_COLS = {
    "type": "A",
    "qty": "B",
    # CO columns: C through AO (dynamic based on header row)
    "co_qty": "AP",
    "po_co_qty": "AQ",
    "thickness": "AR",
    "x": "AS",
    "width": "AT",
    "length": "AU",
    "material_type": "AV",
    "lf_pcs": "AW",
    "bf_sf": "AX",
    "cost_mbf": "AY",
    "total_cost": "AZ",
    "total_cost_tax": "BA",
    "invoice_num": "BB",
    # Delivery date columns: BC through DS (row 2 has dates)
    "total_delivered": "DV",
    "delivered_lf": "DW",
    "delivered_bf_sf": "DX",
    "delivered_cost": "DY",
    "delivered_cost_tax": "DZ",
    "pct_delivery": "EA",
    "inv_bundles": "EB",
    "inv_uom": "EC",
    "pcs_bundle": "ED",
    "inv_pcs": "EE",
    "issues": "EF",
    "issues_lf": "EG",
    "issues_bf_sf": "EH",
    "pct_issued": "EI",
    "issues_cost": "EJ",
    "issues_cost_tax": "EK",
    "variance_code": "EL",
    "reason": "EM",
}

# Column layout for Willow Way sheet  
WILLOW_COLS = {
    "type": "A",
    "qty": "B",
    # CO columns: C through O
    "co_qty": "Q",
    "po_co_qty": "R",
    "thickness": "S",
    "x": "T",
    "width": "U",
    "length": "V",
    "material_type": "W",
    "lf_pcs": "X",
    "bf_sf": "Y",
    "cost_mbf": "Z",
    "total_cost": "AA",
    "total_cost_tax": "AB",
    "invoice_num": "AC",
    # Delivery date columns: AD through AU (row 2 has dates)
    "total_delivered": "AV",
    "delivered_lf": "AW",
    "delivered_bf_sf": "AX",
    "delivered_cost": "AY",
    "delivered_cost_tax": "AZ",
    "pct_delivery": "BA",
    "inv_bundles": "BB",
    "inv_uom": "BC",
    "pcs_bundle": "BD",
    "inv_pcs": "BE",
    "issues": "BF",
    "issues_lf": "BG",
    "issues_bf_sf": "BH",
    "pct_issued": "BI",
    "issues_cost": "BJ",
    "issues_cost_tax": "BK",
    "variance_code": "BL",
    "reason": "BM",
}


def col_to_num(col_letter: str) -> int:
    """Convert column letter to number (1-indexed)."""
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def get_existing_invoice_numbers(ws, invoice_col: str) -> Set[str]:
    """Extract all invoice numbers already recorded in the sheet."""
    col_num = col_to_num(invoice_col)
    existing = set()
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_num)
        if cell.value:
            # Invoice numbers can be in format: "68981028\n68981040" or "68981028 68981040"
            val = str(cell.value)
            for inv_num in re.findall(r'\d{5,}', val):
                existing.add(inv_num)
    return existing


def get_delivery_date_columns(ws, start_col: str, row: int = 2) -> Dict[datetime, int]:
    """Get mapping of delivery dates to column numbers from row 2."""
    date_cols = {}
    start_num = col_to_num(start_col)
    for col in range(start_num, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, datetime):
            date_cols[cell.value] = col
        elif cell.value is None:
            # Check if we've passed the date range
            # Look ahead a few columns
            found_more = False
            for check_col in range(col + 1, min(col + 5, ws.max_column + 1)):
                if isinstance(ws.cell(row=row, column=check_col).value, datetime):
                    found_more = True
                    break
            if not found_more:
                break
    return date_cols


def find_matching_row(ws, invoice_item: LineItem, cols: dict, 
                      data_start_row: int, data_end_row: int) -> Optional[int]:
    """Find which row in the sheet matches this invoice line item.
    
    Matching is based on material description, dimensions, and material type.
    """
    mat_type_col = col_to_num(cols["material_type"])
    thickness_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    length_col = col_to_num(cols["length"])
    type_col = col_to_num(cols["type"])
    
    # Parse dimensions from the invoice item
    inv_desc = invoice_item.description.upper().strip()
    inv_dims = invoice_item.dimensions.strip()
    
    # Extract T, W, L from dimensions like "2X6X12"
    inv_t, inv_w, inv_l = None, None, None
    dm = re.match(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_dims)
    if not dm:
        dm = re.search(r'(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)', inv_desc)
        
    if dm:
        inv_t = int(dm.group(1))
        inv_w = int(dm.group(2))
        inv_l = int(dm.group(3))
    else:
        dm_2d = re.search(r'(\d+)\s*[xX]\s*(\d+)', inv_desc)
        if dm_2d:
            inv_w = int(dm_2d.group(1))
            inv_l = int(dm_2d.group(2))
            
    if not inv_l and "PET 104-5/8" in inv_desc:
        inv_l = 9
    elif not inv_l and "PET 116-5/8" in inv_desc:
        inv_l = 10
    
    # Determine item category
    inv_category = classify_item_category(inv_desc, invoice_item.item_code)
    
    best_match = None
    best_score = 0
    
    for row in range(data_start_row, data_end_row + 1):
        row_type = ws.cell(row=row, column=type_col).value
        if not row_type:
            continue
        row_type = str(row_type).strip()
        
        row_mat = ws.cell(row=row, column=mat_type_col).value
        if row_mat:
            row_mat = str(row_mat).strip().upper()
        else:
            row_mat = ""
        
        row_t = ws.cell(row=row, column=thickness_col).value
        row_w = ws.cell(row=row, column=width_col).value
        row_l = ws.cell(row=row, column=length_col).value
        
        # Convert to comparable values
        try:
            row_t = int(float(str(row_t).strip())) if row_t and str(row_t).strip() not in ('', '\xa0') else None
        except (ValueError, TypeError):
            row_t = None
        try:
            row_w = int(float(str(row_w).strip())) if row_w and str(row_w).strip() not in ('', '\xa0') else None
        except (ValueError, TypeError):
            row_w = None
        try:
            row_l = int(float(str(row_l).strip())) if row_l and str(row_l).strip() not in ('', '\xa0') else None
        except (ValueError, TypeError):
            row_l = None
        
        score = 0
        
        # Category match
        row_category = row_type.lower()
        if inv_category == row_category:
            score += 10
        elif inv_category in row_category or row_category in inv_category:
            score += 5
        else:
            continue  # Skip if category doesn't match at all
        
        # Dimension match
        if inv_t is not None and row_t is not None and inv_t == row_t:
            score += 5
        if inv_w is not None and row_w is not None and inv_w == row_w:
            score += 5
        if inv_l is not None and row_l is not None and inv_l == row_l:
            score += 5
        
        # Material type match
        if row_mat:
            if "PT" in inv_desc and ("PT" in row_mat or "MCA" in row_mat):
                score += 3
            elif "MCA" in inv_desc and ("MCA" in row_mat or "PT" in row_mat):
                score += 3
            elif "SYP" in inv_desc and "SYP" in row_mat:
                score += 3
            elif "LVL" in inv_desc and "LVL" in row_mat:
                score += 3
            elif "OSB" in inv_desc and "OSB" in row_mat:
                score += 3
            elif "PLY" in inv_desc and "PLY" in row_mat:
                score += 3
            elif "ZIP" in inv_desc and "ZIP" in row_mat:
                score += 3
            # "Each" items - match by description keywords
            if row_type.lower() in ("each", "invoice"):
                desc_words = set(row_mat.split())
                inv_words = set(inv_desc.split())
                common = desc_words & inv_words
                score += len(common) * 2
        
        if score > best_score:
            best_score = score
            best_match = row
    
    # Only return a match if score is high enough
    if best_score >= 15:
        return best_match
    return None


def classify_item_category(desc: str, item_code: str = "") -> str:
    """Classify a material item into a category."""
    desc = desc.upper()
    code = item_code.upper()
    
    if any(k in desc for k in ["LVL", "PSL", "GLB", "GLULAM", "LSL"]):
        return "lvl"
    if any(k in desc for k in ["OSB", "PLYWOOD", "PLY", "ZIP", "CDX", "GYPSUM", "SHEATHING"]):
        return "panels"
    if any(k in desc for k in ["SILL SEAL", "ADHESIVE", "TAPE", "FLASHING", "TYVEK", 
                                "WRAP", "SEALANT", "CAULK", "DYNAFLEX", "THRESHOLD",
                                "BT20", "COMPOUND"]):
        return "each"
    if any(k in desc for k in ["SYP", "MCA", "TREATED", "LUMBER"]):
        return "lumber"
    if any(k in code for k in ["SYP", "MCA"]):
        return "lumber"
    if any(k in code for k in ["LVL", "PSL"]):
        return "lvl"
    if any(k in code for k in ["OSB", "PLY", "ZIP"]):
        return "panels"
    if any(k in code for k in ["SILL", "TAPE", "TYVEK", "CAULK", "ADHES"]):
        return "each"
    
    return "lumber"  # Default


def find_or_add_delivery_date_col(ws, ship_date: datetime, 
                                   date_cols: Dict[datetime, int],
                                   date_row: int = 2) -> int:
    """Find the column for this delivery date, or add a new one."""
    # Normalize to date only (no time)
    ship_date_norm = datetime(ship_date.year, ship_date.month, ship_date.day)
    
    if ship_date_norm in date_cols:
        return date_cols[ship_date_norm]
    
    # Find the closest date and insert after it or at the end
    if date_cols:
        sorted_dates = sorted(date_cols.keys())
        # Find insertion point
        insert_after = None
        for d in sorted_dates:
            if d < ship_date_norm:
                insert_after = d
            else:
                break
        
        if insert_after:
            col = date_cols[insert_after] + 1
        else:
            col = date_cols[sorted_dates[0]]  # Before the first date
        
        # We won't insert columns (too complex with formula references)
        # Instead, find the next empty date column
        max_date_col = max(date_cols.values())
        new_col = max_date_col + 1
        
        # Skip over the rest of the previous date's merged span
        while type(ws.cell(row=date_row, column=new_col)).__name__ == 'MergedCell':
            new_col += 1
            
        # Set the date in row 2
        ws.cell(row=date_row, column=new_col, value=ship_date_norm)
        date_cols[ship_date_norm] = new_col
        return new_col
    
    return None


def update_invoice_reference(ws, row: int, col_num: int, invoice_number: str):
    """Add an invoice number to the invoice reference cell (appending if needed)."""
    cell = ws.cell(row=row, column=col_num)
    current = cell.value
    
    if current is None:
        cell.value = invoice_number
    else:
        current_str = str(current)
        # Check if already there
        if invoice_number in current_str:
            return
        # Append with newline
        cell.value = current_str + "\n" + invoice_number


def process_sheet(ws, sheet_name: str, invoices: List[InvoiceData], 
                  cols: dict, project_name: str,
                  delivery_start_col: str, date_row: int = 2,
                  data_ranges: List[Tuple[int, int]] = None):
    """Process a sheet: match invoices to rows and update delivery data."""
    
    print(f"\n{'='*60}")
    print(f"Processing sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # Get existing invoice numbers
    inv_col_num = col_to_num(cols["invoice_num"])
    existing_invs = get_existing_invoice_numbers(ws, cols["invoice_num"])
    print(f"  Existing invoice references: {len(existing_invs)}")
    
    # Get delivery date columns
    date_cols = get_delivery_date_columns(ws, delivery_start_col, date_row)
    print(f"  Existing delivery date columns: {len(date_cols)}")
    
    # Filter invoices for this project
    project_invoices = [inv for inv in invoices 
                        if inv.project_name == project_name]
    print(f"  Total invoices for {project_name}: {len(project_invoices)}")
    
    # Find new invoices
    new_invoices = [inv for inv in project_invoices 
                    if inv.invoice_number not in existing_invs]
    print(f"  New invoices to process: {len(new_invoices)}")
    
    if not data_ranges:
        data_ranges = [(3, ws.max_row)]
    
    matched_count = 0
    unmatched_items = []
    
    for inv in new_invoices:
        if not inv.ship_date:
            print(f"  WARNING: Invoice {inv.invoice_number} has no ship date, skipping")
            continue
            
        for item in inv.line_items:
            # Try to find matching row
            matched_row = None
            for start_row, end_row in data_ranges:
                matched_row = find_matching_row(ws, item, cols, start_row, end_row)
                if matched_row:
                    break
            
            if matched_row:
                matched_count += 1
                
                # Add invoice number reference
                update_invoice_reference(ws, matched_row, inv_col_num, inv.invoice_number)
                
                # Add delivery quantity to the appropriate date column
                date_col = find_or_add_delivery_date_col(ws, inv.ship_date, date_cols, date_row)
                if date_col:
                    # Add the quantity (in pieces or the appropriate unit)
                    current_val = ws.cell(row=matched_row, column=date_col).value
                    qty = item.quantity
                    if current_val and isinstance(current_val, (int, float)):
                        qty += current_val
                    elif current_val and isinstance(current_val, str):
                        # It might be a formula - append
                        try:
                            qty += float(current_val)
                        except ValueError:
                            pass  # Keep the new qty
                    ws.cell(row=matched_row, column=date_col, value=qty)
            else:
                unmatched_items.append((inv.invoice_number, item))
    
    print(f"  Matched items: {matched_count}")
    print(f"  Unmatched items: {len(unmatched_items)}")
    
    if unmatched_items:
        print(f"\n  --- Unmatched Items ---")
        for inv_num, item in unmatched_items[:20]:
            print(f"    Invoice {inv_num}: {item.quantity} {item.uom} "
                  f"- {item.description} ({item.dimensions})")
        if len(unmatched_items) > 20:
            print(f"    ... and {len(unmatched_items) - 20} more")
    
    # Update total delivered formulas and inventory columns
    update_delivery_totals(ws, cols, delivery_start_col, date_cols, data_ranges)
    
    return matched_count, unmatched_items


def update_delivery_totals(ws, cols: dict, delivery_start_col: str,
                           date_cols: Dict[datetime, int],
                           data_ranges: List[Tuple[int, int]]):
    """Update the Total Delivered and downstream formulas for all data rows."""
    
    if not date_cols:
        return
    
    del_start = min(date_cols.values())
    del_end = max(date_cols.values())
    
    total_del_col = col_to_num(cols["total_delivered"])
    del_lf_col = col_to_num(cols["delivered_lf"])
    del_bf_col = col_to_num(cols["delivered_bf_sf"])
    del_cost_col = col_to_num(cols["delivered_cost"])
    del_cost_tax_col = col_to_num(cols["delivered_cost_tax"])
    pct_del_col = col_to_num(cols["pct_delivery"])
    
    length_col = col_to_num(cols["length"])
    thickness_col = col_to_num(cols["thickness"])
    width_col = col_to_num(cols["width"])
    cost_col = col_to_num(cols["cost_mbf"])
    total_cost_col = col_to_num(cols["total_cost"])
    type_col = col_to_num(cols["type"])
    
    # Inventory columns
    inv_bundles_col = col_to_num(cols["inv_bundles"])
    inv_uom_col = col_to_num(cols["inv_uom"])
    pcs_bundle_col = col_to_num(cols["pcs_bundle"])
    inv_pcs_col = col_to_num(cols["inv_pcs"])
    issues_col = col_to_num(cols["issues"])
    issues_lf_col = col_to_num(cols["issues_lf"])
    issues_bf_col = col_to_num(cols["issues_bf_sf"])
    pct_issued_col = col_to_num(cols["pct_issued"])
    issues_cost_col = col_to_num(cols["issues_cost"])
    issues_cost_tax_col = col_to_num(cols["issues_cost_tax"])
    
    del_start_letter = get_column_letter(del_start)
    del_end_letter = get_column_letter(del_end)
    
    for start_row, end_row in data_ranges:
        for row in range(start_row, end_row + 1):
            row_type = ws.cell(row=row, column=type_col).value
            if not row_type:
                continue
            
            row_str = str(row)
            
            # Total Delivered = SUM(delivery date columns)
            total_del_formula = f"=SUM({del_start_letter}{row}:{del_end_letter}{row})"
            ws.cell(row=row, column=total_del_col, value=total_del_formula)
            
            # Get column letters for formulas
            td_letter = get_column_letter(total_del_col)
            l_letter = get_column_letter(length_col)
            t_letter = get_column_letter(thickness_col)
            w_letter = get_column_letter(width_col)
            cost_letter = get_column_letter(cost_col)
            tc_letter = get_column_letter(total_cost_col)
            
            # Delivered L/F = Total Delivered × Length
            ws.cell(row=row, column=del_lf_col, 
                    value=f"={td_letter}{row}*{l_letter}{row}")
            
            # Check if this is a lumber/LVL item (has thickness/width) or Each/Panel
            row_type_val = str(row_type).strip().lower()
            
            if row_type_val in ("lumber",):
                # Delivered B/F = Total Delivered × T × W × L / 12
                ws.cell(row=row, column=del_bf_col,
                        value=f"={td_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row}/12")
                # Delivered Cost = (B/F × Cost/MBF) / 1000
                bf_letter = get_column_letter(del_bf_col)
                ws.cell(row=row, column=del_cost_col,
                        value=f"={bf_letter}{row}*{cost_letter}{row}/1000")
            elif row_type_val == "panels":
                # Panels: SF-based
                ws.cell(row=row, column=del_bf_col,
                        value=f"={td_letter}{row}*{t_letter}{row}*{w_letter}{row}")
                bf_letter = get_column_letter(del_bf_col)
                ws.cell(row=row, column=del_cost_col,
                        value=f"={bf_letter}{row}*{cost_letter}{row}/1000")
            elif row_type_val in ("lvl",):
                # LVL: LF-based pricing
                lf_letter = get_column_letter(del_lf_col)
                ws.cell(row=row, column=del_cost_col,
                        value=f"={lf_letter}{row}*{cost_letter}{row}")
            elif row_type_val in ("each", "invoice"):
                # Each items: simple qty × price
                ws.cell(row=row, column=del_cost_col,
                        value=f"={td_letter}{row}*{cost_letter}{row}")
            
            # Delivered Cost with Tax = Cost × 1.06
            dc_letter = get_column_letter(del_cost_col)
            ws.cell(row=row, column=del_cost_tax_col,
                    value=f"={dc_letter}{row}*1.06")
            
            # % Delivery = Delivered Cost / Total Cost
            ws.cell(row=row, column=pct_del_col,
                    value=f"=IFERROR({dc_letter}{row}/{tc_letter}{row},0)")
            
            # --- Inventory Automation ---
            # Inventory in PCS = Inventory Bundles × PCS/Bundle (if bundles tracked)
            ib_letter = get_column_letter(inv_bundles_col)
            pb_letter = get_column_letter(pcs_bundle_col)
            ip_letter = get_column_letter(inv_pcs_col)
            ws.cell(row=row, column=inv_pcs_col,
                    value=f"=IF({ib_letter}{row}<>\"\",{ib_letter}{row}*{pb_letter}{row},0)")
            
            # Issues = Total Delivered - Inventory PCS (what's been used)
            ws.cell(row=row, column=issues_col,
                    value=f"={td_letter}{row}-{ip_letter}{row}")
            
            # Issues L/F
            iss_letter = get_column_letter(issues_col)
            ws.cell(row=row, column=issues_lf_col,
                    value=f"={iss_letter}{row}*{l_letter}{row}")
            
            # Issues B/F-S/F
            if row_type_val in ("lumber",):
                ws.cell(row=row, column=issues_bf_col,
                        value=f"=({iss_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row})/12")
            elif row_type_val == "panels":
                ws.cell(row=row, column=issues_bf_col,
                        value=f"={iss_letter}{row}*{t_letter}{row}*{w_letter}{row}")
            else:
                ws.cell(row=row, column=issues_bf_col,
                        value=f"={iss_letter}{row}")
            
            # % Issued
            ibf_letter = get_column_letter(issues_bf_col)
            ws.cell(row=row, column=pct_issued_col,
                    value=f"=IFERROR({ibf_letter}{row}/{td_letter}{row},0)")
            
            # Issues Cost
            ws.cell(row=row, column=issues_cost_col,
                    value=f"=IFERROR({ibf_letter}{row}*{cost_letter}{row}/1000,0)")
            
            # Issues Cost with Tax
            ic_letter = get_column_letter(issues_cost_col)
            ws.cell(row=row, column=issues_cost_tax_col,
                    value=f"={ic_letter}{row}*1.06")


def safe_set_cell(ws, row, col, val):
    try:
        ws.cell(row=row, column=col, value=val)
    except AttributeError:
        pass # Ignore MergedCell error

def seed_sheet_from_pos(ws, sheet_name: str, pos: List, cols: dict, project_name: str):
    print(f"\n{'='*60}")
    print(f"Seeding sheet: {sheet_name} from POs")
    print(f"{'='*60}")
    
    project_pos = [po for po in pos if project_name.lower() in po.project_name.lower() or po.project_name == "Unknown"]
    print(f"  Found {len(project_pos)} POs for {project_name}")
    
    # Find the last empty row
    type_col = col_to_num(cols["type"])
    
    # For a completely blank template, we start at row 3
    # Check if it's actually blank
    is_blank = True
    for r in range(3, 10):
        if ws.cell(row=r, column=type_col).value:
            is_blank = False
            break
            
    if not is_blank:
        print("  Sheet already has data. Skipping PO seeding.")
        return
        
    print("  Sheet is blank. Generating rows from POs...")
    
    qty_col = col_to_num(cols["qty"])
    po_co_qty_col = col_to_num(cols["po_co_qty"])
    t_col = col_to_num(cols["thickness"])
    w_col = col_to_num(cols["width"])
    l_col = col_to_num(cols["length"])
    mat_type_col = col_to_num(cols["material_type"])
    cost_col = col_to_num(cols["cost_mbf"])
    total_cost_col = col_to_num(cols["total_cost"])
    
    row = 3
    added_items = 0
    
    for po in project_pos:
        for item in po.line_items:
            category = item.category.capitalize() if hasattr(item, 'category') else "Lumber"
            safe_set_cell(ws, row, type_col, category)
            
            # Initial QTY equals PO qty
            safe_set_cell(ws, row, qty_col, item.quantity)
            safe_set_cell(ws, row, po_co_qty_col, item.quantity)
            
            # Dimensions
            if hasattr(item, 'dimensions') and item.dimensions:
                dm = re.match(r'(\d+)X(\d+)X(\d+)', item.dimensions)
                if dm:
                    safe_set_cell(ws, row, t_col, int(dm.group(1)))
                    safe_set_cell(ws, row, w_col, int(dm.group(2)))
                    safe_set_cell(ws, row, l_col, int(dm.group(3)))
            
            # Description & Price
            safe_set_cell(ws, row, mat_type_col, item.description)
            safe_set_cell(ws, row, cost_col, item.unit_price)
            
            # Total Cost Formula
            qty_letter = get_column_letter(qty_col)
            cost_letter = get_column_letter(cost_col)
            if category.lower() in ["lumber"]:
                t_letter = get_column_letter(t_col)
                w_letter = get_column_letter(w_col)
                l_letter = get_column_letter(l_col)
                safe_set_cell(ws, row, total_cost_col, f"=({qty_letter}{row}*{t_letter}{row}*{w_letter}{row}*{l_letter}{row}/12)*{cost_letter}{row}/1000")
            elif category.lower() in ["panels"]:
                t_letter = get_column_letter(t_col)
                w_letter = get_column_letter(w_col)
                safe_set_cell(ws, row, total_cost_col, f"=({qty_letter}{row}*{t_letter}{row}*{w_letter}{row})*{cost_letter}{row}/1000")
            else:
                safe_set_cell(ws, row, total_cost_col, f"={qty_letter}{row}*{cost_letter}{row}")
            
            row += 1
            added_items += 1
            
    print(f"  Seeded {added_items} items into {sheet_name} from POs.")


def process_vpos_sheet(ws, change_orders: List[ChangeOrderData]):
    """Update the VPO's sheet with any new change orders."""
    print(f"\n{'='*60}")
    print(f"Processing VPO's sheet")
    print(f"{'='*60}")
    
    # Get existing CO numbers in the sheet
    existing_cos = set()
    for row in range(1, ws.max_row + 1):
        cell_k = ws.cell(row=row, column=11)  # Column K = CO numbers
        if cell_k.value:
            existing_cos.add(str(cell_k.value).strip())
    
    cobia_cos = [co for co in change_orders if co.project_name == "Cobia Cove"]
    willow_cos = [co for co in change_orders if co.project_name == "Willow Way"]
    
    print(f"  Total COs: Cobia={len(cobia_cos)}, Willow={len(willow_cos)}")
    print(f"  Existing CO references in sheet: {len(existing_cos)}")
    
    # For now, we won't modify the VPOs sheet structure as it has a specific
    # manual layout. We just report what's available.
    print(f"  VPO sheet preserved as-is (manual CO tracking)")


def main():
    """Main automation pipeline."""
    print("=" * 60)
    print("KNCC Material Procurement Automation")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    dry_run = "--dry-run" in sys.argv
    if dry_run:
        print("*** DRY RUN MODE - no changes will be saved ***\n")
    
    # Step 1: Parse all PDFs
    print("\n[Step 1] Scanning and parsing all PDFs...")
    data = scan_and_parse_all(BASE_PATH)
    
    print(f"  Invoices parsed: {len(data['invoices'])}")
    print(f"  Change Orders parsed: {len(data['change_orders'])}")
    print(f"  Purchase Orders parsed: {len(data['purchase_orders'])}")
    if data["warnings"]:
        for w in data["warnings"]:
            print(f"  WARNING: {w}")
    
    # Step 2: Load existing Excel
    print(f"\n[Step 2] Loading existing Excel: {INPUT_EXCEL}")
    if not os.path.exists(INPUT_EXCEL):
        print(f"  ERROR: File not found: {INPUT_EXCEL}")
        return
    
    # Create backup
    if not dry_run:
        import shutil
        shutil.copy2(INPUT_EXCEL, BACKUP_EXCEL)
        print(f"  Backup created: {BACKUP_EXCEL}")
    
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    print(f"  Sheets: {wb.sheetnames}")
    
    # Step 2.5: Seed from POs
    print(f"\n[Step 2.5] Seeding blank template from Purchase Orders...")
    if SHEET_COBIA in wb.sheetnames:
        seed_sheet_from_pos(wb[SHEET_COBIA], SHEET_COBIA, data["purchase_orders"], COBIA_COLS, "Cobia Cove")
    if SHEET_WILLOW in wb.sheetnames:
        seed_sheet_from_pos(wb[SHEET_WILLOW], SHEET_WILLOW, data["purchase_orders"], WILLOW_COLS, "Willow Way")
    
    # Step 3: Process Cobia Cove sheet
    if SHEET_COBIA in wb.sheetnames:
        ws_cobia = wb[SHEET_COBIA]
        # Data ranges for Cobia Cove:
        # Rows 3-118: Main lumber/materials
        # Rows 123-152: Second section
        # Rows 157-170: House wrap
        # Rows 173-176: Mockup
        cobia_ranges = [(3, 118), (123, 152), (157, 170), (173, 176)]
        
        matched, unmatched = process_sheet(
            ws_cobia, SHEET_COBIA, data["invoices"],
            COBIA_COLS, "Cobia Cove", "BC", 2, cobia_ranges
        )
    
    # Step 4: Process Willow Way sheet
    if SHEET_WILLOW in wb.sheetnames:
        ws_willow = wb[SHEET_WILLOW]
        # Data ranges for Willow Way:
        # Rows 3-78: All data
        willow_ranges = [(3, 78)]
        
        matched, unmatched = process_sheet(
            ws_willow, SHEET_WILLOW, data["invoices"],
            WILLOW_COLS, "Willow Way", "AD", 2, willow_ranges
        )
    
    # Step 5: Process VPO's sheet
    if SHEET_VPOS in wb.sheetnames:
        ws_vpos = wb[SHEET_VPOS]
        process_vpos_sheet(ws_vpos, data["change_orders"])
    
    # Step 6: Save
    if not dry_run:
        print(f"\n[Step 6] Saving to: {OUTPUT_EXCEL}")
        wb.save(OUTPUT_EXCEL)
        print("  ✓ Saved successfully!")
    else:
        print(f"\n[Step 6] DRY RUN - not saving")
    
    print(f"\n{'='*60}")
    print(f"Automation complete: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
