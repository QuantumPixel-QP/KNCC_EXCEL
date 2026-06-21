"""
Quick verification that the Excel was updated correctly.
Check some key cells and formulas.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('Client/Client_Requirments_Doc.xlsx')

# Check Cobia Cove sheet
ws = wb['Cobia Cove Appartments']
print("=== Cobia Cove Appartments ===")
print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} cols")

# Check that delivery total formulas exist
print("\n  Sample delivery total formulas (col DV):")
dv_col = 126  # DV
for row in [3, 5, 10, 50, 100, 150]:
    cell = ws.cell(row=row, column=dv_col)
    if cell.value:
        print(f"    Row {row}: {cell.value}")

# Check inventory formulas
print("\n  Sample inventory formulas:")
for row in [3, 5, 10]:
    ee_cell = ws.cell(row=row, column=135)  # EE = Inventory in PCS
    ef_cell = ws.cell(row=row, column=136)  # EF = Issues
    ei_cell = ws.cell(row=row, column=139)  # EI = % Issued
    print(f"    Row {row}: InvPCS={ee_cell.value}, Issues={ef_cell.value}, %Issued={ei_cell.value}")

# Check % Delivery formulas
print("\n  Sample % Delivery formulas (col EA):")
ea_col = 131  # EA
for row in [3, 5, 10, 50]:
    cell = ws.cell(row=row, column=ea_col)
    if cell.value:
        print(f"    Row {row}: {cell.value}")

# Check Willow Way sheet
ws2 = wb['Willow Way Apts']
print("\n=== Willow Way Apts ===")
print(f"  Dimensions: {ws2.max_row} rows x {ws2.max_column} cols")

# Check delivery formulas
av_col = 48  # AV = Total Delivered
for row in [3, 5, 10, 20]:
    cell = ws2.cell(row=row, column=av_col)
    if cell.value:
        print(f"    Row {row} Total Del: {cell.value}")

# Check new delivery date column
print("\n  Delivery date columns:")
for col in range(30, ws2.max_column + 1):
    cell = ws2.cell(row=2, column=col)
    if cell.value:
        from openpyxl.utils import get_column_letter
        print(f"    {get_column_letter(col)}{2}: {cell.value}")

# Check that invoice ref for matched item was added (60126038)
ac_col = 29  # AC = Invoice #
print("\n  Invoice references check:")
for row in range(3, 20):
    cell = ws2.cell(row=row, column=ac_col)
    mat_cell = ws2.cell(row=row, column=23)  # W = Material Type
    if cell.value:
        print(f"    Row {row}: {cell.value} | Material: {mat_cell.value}")

wb.close()
