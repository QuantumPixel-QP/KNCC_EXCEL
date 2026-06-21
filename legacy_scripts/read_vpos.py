import openpyxl
wb = openpyxl.load_workbook('Client/Client_Requirments_Doc.xlsx')
ws = wb["VPO's"]
print(f"Sheet: VPOs (rows={ws.max_row}, cols={ws.max_column})")
for row_num in range(1, min(20, ws.max_row+1)):
    cells = []
    for cell in ws[row_num]:
        if cell.value is not None:
            cells.append(f"{cell.column_letter}:{repr(cell.value)}")
    if cells:
        print(f"  Row {row_num}: {cells}")
    else:
        print(f"  Row {row_num}: (empty)")
