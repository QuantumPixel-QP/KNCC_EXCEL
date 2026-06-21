"""
compare_excel.py — Cell-by-cell diff between generated and original Excel.
Checks: values in key data columns, formula presence in formula columns.
Outputs a summary of matches and mismatches.
"""
import openpyxl
from openpyxl.utils import get_column_letter

ORIG = r'C:\Users\Admin\Desktop\KNCC_EXCEL\Client\Client_Requirments_Doc.xlsx'
GEN  = r'C:\Users\Admin\Desktop\KNCC_EXCEL\platform\backend\exports\Willow_Way_Village_Requirements.xlsx'
SHEET = 'Willow Way Apts'

# Columns to compare data values (from WILLOW_COLS)
DATA_COLS = {
    'A':  'type',
    'B':  'qty',
    'Q':  'co_qty',
    'R':  'po_co_qty',
    'S':  'thickness',
    'U':  'width',
    'V':  'length',
    'W':  'material_type',
    'X':  'lf_pcs',
    'Y':  'bf_sf',
    'Z':  'cost_mbf',
    'AA': 'total_cost',
    'AB': 'total_cost_tax',
    'AC': 'invoice_num',
}

# Columns that should have formulas in the generated file
FORMULA_COLS = ['AV','AW','AX','AY','AZ','BA','BE','BF','BG','BH','BI','BJ','BK']

DATA_ROWS = list(range(3, 79))

def col_to_num(c):
    n = 0
    for ch in c.upper():
        n = n * 26 + (ord(ch) - 64)
    return n

def cell_val(ws, row, col_letter):
    return ws.cell(row=row, column=col_to_num(col_letter)).value

orig_wb = openpyxl.load_workbook(ORIG, data_only=True)
gen_wb  = openpyxl.load_workbook(GEN,  data_only=False)  # keep formulas

if SHEET not in orig_wb.sheetnames:
    print(f'ERROR: Sheet "{SHEET}" not in original. Available: {orig_wb.sheetnames}')
    exit(1)
if SHEET not in gen_wb.sheetnames:
    print(f'ERROR: Sheet "{SHEET}" not in generated. Available: {gen_wb.sheetnames}')
    exit(1)

orig_ws = orig_wb[SHEET]
gen_ws  = gen_wb[SHEET]

# ── DATA value comparison ─────────────────────────────────────
total_checked = 0
mismatches = []
empty_orig  = 0

for row in DATA_ROWS:
    orig_type = cell_val(orig_ws, row, 'A')
    if not orig_type:
        empty_orig += 1
        continue

    for col_letter, field_name in DATA_COLS.items():
        orig_val = cell_val(orig_ws, row, col_letter)
        gen_val  = cell_val(gen_ws, row, col_letter)
        total_checked += 1

        # Normalise: None == 0 for numeric fields
        def norm(v):
            if v is None or v == '' or v == '-':
                return None
            try:
                f = float(str(v).replace(',',''))
                return round(f, 2)
            except:
                return str(v).strip()

        ov = norm(orig_val)
        gv = norm(gen_val)

        if ov != gv:
            mismatches.append({
                'row': row,
                'col': col_letter,
                'field': field_name,
                'orig': orig_val,
                'gen': gen_val,
            })

print(f'\n=== DATA COMPARISON (rows 3-78, {len(DATA_COLS)} cols) ===')
print(f'  Rows with data:  {len(DATA_ROWS) - empty_orig}')
print(f'  Cells checked:   {total_checked}')
print(f'  Mismatches:      {len(mismatches)}')
match_pct = (total_checked - len(mismatches)) / total_checked * 100 if total_checked else 0
print(f'  Match rate:      {match_pct:.1f}%')

if mismatches:
    print('\n  First 20 mismatches:')
    for m in mismatches[:20]:
        print(f'    Row {m["row"]:3d} Col {m["col"]:3s} [{m["field"]:20s}] orig={repr(m["orig"]):<25} gen={repr(m["gen"])}')

# ── FORMULA presence check ────────────────────────────────────
print(f'\n=== FORMULA CHECK (generated file) ===')
formula_ok = 0
formula_missing = []
for row in DATA_ROWS:
    gen_type = cell_val(gen_ws, row, 'A')  # Note: data_only=False so formulas show
    if not gen_type:
        continue
    for col in FORMULA_COLS:
        raw = gen_ws.cell(row=row, column=col_to_num(col)).value
        if raw and str(raw).startswith('='):
            formula_ok += 1
        else:
            formula_missing.append((row, col))

print(f'  Formula cells present: {formula_ok}')
print(f'  Formula cells missing: {len(formula_missing)}')
if formula_missing[:10]:
    print(f'  Missing examples: {formula_missing[:10]}')

# ── Delivery date columns ─────────────────────────────────────
print(f'\n=== DELIVERY DATE COLUMNS (row 2) ===')
orig_dates = []
gen_dates  = []
for col in range(col_to_num('AD'), col_to_num('AV')):
    ov = orig_ws.cell(row=2, column=col).value
    gv = gen_ws.cell(row=2, column=col).value
    if ov: orig_dates.append(str(ov)[:10])
    if gv: gen_dates.append(str(gv)[:10])

print(f'  Orig delivery dates: {len(orig_dates)}  → {orig_dates[:5]}')
print(f'  Gen  delivery dates: {len(gen_dates)}   → {gen_dates[:5]}')

print('\n=== SUMMARY ===')
if match_pct >= 95 and len(formula_missing) < 10:
    print('✅ EXCELLENT — Generated Excel closely matches the original!')
elif match_pct >= 80:
    print('⚠️  GOOD — Minor data discrepancies found. Review mismatches above.')
else:
    print('❌ NEEDS WORK — Significant discrepancies. Check import/export pipeline.')
