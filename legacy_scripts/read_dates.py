import openpyxl
wb = openpyxl.load_workbook('Client/Client_Requirments_Doc.xlsx')

ws = wb['Cobia Cove Appartments']
# Row 2 - delivery dates  
print("=== Row 2 (Delivery Dates) ===")
for cell in ws[2]:
    if cell.value is not None:
        print(f"  {cell.coordinate} ({cell.column_letter}, col {cell.column}): {repr(cell.value)}")

# Row 1 headers in column range BB-EN
print("\n=== Row 1 Headers (BB onward) ===")
for cell in ws[1]:
    if cell.value is not None and cell.column >= 54:  # BB = 54
        print(f"  {cell.coordinate} ({cell.column_letter}, col {cell.column}): {repr(cell.value)}")
